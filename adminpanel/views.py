import openpyxl
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render,redirect
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Sum, Count, Case, When, Value, Q
from teacher.models import TeacherInfo
from student.models import StudentInfo
from django.contrib.auth.decorators import login_required
from users.decorators import admin_required
from .models import *
import json

def login(request):
    return render(request, 'adminpanel/login.html')


@login_required
@admin_required
def index(request):
    course = Courses.objects.all().order_by('level', 'term', 'courseCode')
    student = StudentInfo.objects.all()
    teacher = TeacherInfo.objects.all().values('tID','id','tInitial','tPhone','tEmail','tName','tDesignation').annotate(numofstudent=Count('studentinfo',distinct=True))
    semester = SemesterInfo.objects.all()
    regStudents= CoursePreRegistration.objects.filter().values('student_id').distinct()
    waivers = WaiverInfo.objects.filter(semester__semesterCode='201')
    finalregcount = CourseRegistration.objects.filter(registered=True).count()
    context = {
        'courses':course,
        'teachers':teacher,
        'students':student,
        'semesters':semester,
        'regStudents':regStudents,
        'waivers':waivers,
        'finalregcount':finalregcount,
    }
    return render(request, 'adminpanel/index.html',context)
def insert(request):
    try:
        table = request.POST['editorTitle']
        if table == 'course':
            courseCode = request.POST['ccode']
            courseTitle = request.POST['ctitle']
            courseCredit = request.POST['ccredit']
            courseLevel = request.POST['clevel']
            courseTerm = request.POST['cterm']
            totalSection = request.POST['tsection']
            if Courses.objects.filter(courseCode=courseCode):
                course = Courses.objects.get(courseCode=courseCode)
                course.courseTitle = courseTitle
                course.courseCredit = courseCredit
                course.level = courseLevel
                course.term = courseTerm
                course.totalSection = totalSection

                course.save()
            else:
                newCourse = Courses(courseCode=courseCode,courseTitle=courseTitle,courseCredit=courseCredit,level=courseLevel,term = courseTerm ,totalSection=totalSection)
                newCourse.save()

            return redirect('index')

        elif table == 'student':
            studentID = request.POST['stid']
            studentName = request.POST['stname']
            studentEmail = request.POST['stemail']
            studentGender = request.POST['stgender']
            studentPhone = request.POST['stphone']
            studentAdvisor = request.POST['stadvisor']
            print(studentAdvisor)
            advisor = TeacherInfo.objects.get(tInitial=studentAdvisor)

            if StudentInfo.objects.filter(stID=studentID).exists():
                student = StudentInfo.objects.get(stID=studentID)
                student.stName = studentName
                student.stEmail = studentEmail
                student.stGender = studentGender
                student.stPhone = studentPhone
                student.stAdvisor = advisor
                student.save()
            else:
                newStudent = StudentInfo(stID=studentID,stName=studentName,stEmail=studentEmail,stGender=studentGender,stPhone=studentPhone,stAdvisor=advisor)
                newStudent.save()

            return redirect('index')

        elif table =='teacher':
            print(table)
            teacherID = request.POST['tid']
            teacherName = request.POST['tname']
            teacherInitial = request.POST['tinitial']
            teacherDesignation = request.POST['tdesingnation']
            teacherPhone = request.POST['tphone']
            teacherEmail = request.POST['temail']
            if TeacherInfo.objects.filter(tID= teacherID).exists():
                teacher = TeacherInfo.objects.get(tID=teacherID)
                teacher.tName = teacherName
                teacher.tInitial = teacherInitial
                teacher.tDesignation = teacherDesignation
                teacher.tPhone = teacherPhone
                teacher.tEmail = teacherEmail
                teacher.save()
            else:
                newTeacher = TeacherInfo(tID=teacherID,tName=teacherName,tInitial=teacherInitial,tDesignation=teacherDesignation,tPhone=teacherPhone,tEmail=teacherEmail)
                newTeacher.save()

            return redirect('index')
        elif table =='semester':
            semesterCode = request.POST['scode']
            semesterTitle = request.POST['stitle']
            regOpenDate = request.POST['regOpenDate']
            # formatedregOpenDate = regOpenDate.strftime("%YYYY-%MM-%DD")
            regClosedDate = request.POST['regClosedDate']
            # formatedregClosedDate = regClosedDate.strftime("%YYYY-%MM-%DD")
            print(regOpenDate)

            if SemesterInfo.objects.filter(semesterCode=semesterCode).exists():
                semester = SemesterInfo.objects.get(semesterCode=semesterCode)
                semester.semesterCode = semesterCode
                semester.semesterTitle = semesterTitle
                semester.regOpenDate = regOpenDate
                semester.regClosedDate = regClosedDate
                semester.save()
            else:
                newSemester = SemesterInfo(semesterCode=semesterCode,semesterTitle=semesterTitle,regOpenDate=regOpenDate,regCloseDate=regClosedDate)
                newSemester.save()

            messages.success(request, f'successfully insert into the system!')
            return redirect('index')
    except :
        messages.error(request, f'Something went wrong. Please insert input properly!')
        return redirect(index)
        # return redirect('error404','Something wromh!!!')




def teacherDelete(request,tid):
    try:
        teacher = TeacherInfo.objects.get(pk=tid)
        teacher.delete()
        messages.success(request, f'Successfully item deleted!')
        return redirect('index')
    except:
        messages.error(request, f'Something went wrong. Operation failed!')
        return redirect('index')


def studentDelete(request,stid):
    try:
        student = StudentInfo.objects.get(pk=stid)
        student.delete()
        messages.success(request, f'Successfully item deleted!')
        return redirect('index')
    except:
        messages.error(request, f'Something went wrong. Operation failed!')
        return redirect('index')


def courseDelete(request,courseCode):
    try:
        course = Courses.objects.get(pk=courseCode)
        course.delete()
        messages.success(request, f'Successfully item deleted!')
        return redirect('index')
    except:
        messages.error(request, f'Something went wrong. Operation failed!')
        return redirect('index')


def studentList(request, courseCode, semesterID):

    # student = StudentInfo.objects.get(stEmail=request.user.email)
    # semesterCode = request.POST['semester']
    # courseCode = request.POST['ccode']
    semester = SemesterInfo.objects.get(semesterCode=semesterID)
    course = Courses.objects.get(id=courseCode)
    studentlistA = CoursePreRegistration.objects.filter(course=course,semester=semester,section='A')
    studentlistB = CoursePreRegistration.objects.filter(course=course,semester=semester,section='B')
    studentlistC = CoursePreRegistration.objects.filter(course=course,semester=semester,section='C')
    studentlistD = CoursePreRegistration.objects.filter(course=course,semester=semester,section='D')
    studentlistE = CoursePreRegistration.objects.filter(course=course,semester=semester,section='E')
    studentlistF = CoursePreRegistration.objects.filter(course=course,semester=semester,section='F')
    context = {
        'course':course,
        'studentlistsA': studentlistA,
        'studentlistsB': studentlistB,
        'studentlistsC': studentlistC,
        'studentlistsD': studentlistD,
        'studentlistsE': studentlistE,
        'studentlistsF': studentlistF,
    }
    return render(request, 'adminpanel/studentlist.html',context)


# def fupload(request):
#
#     if "GET" == request.method:
#         return render(request, 'adminpanel/index.html', {})
#     else:
#         doc = request.FILES
#         excel_file = doc['excel_file']
#         if (str(excel_file).split(‘.’)[-1] == “xls”):
#             data = xls_get(excel_file, column_limit=4)
#         elif (str(excel_file).split(‘.’)[-1] == “xlsx”):
#             data = xlsx_get(excel_file, column_limit=4)
#         else:
#             return redirect( < your_upload_file_fail_url >)
@csrf_exempt
def fileUpload(request):
    try:
        if "GET" == request.method:
            return render(request, 'adminpanel/index.html', {})
        else:
            doc = request.FILES
            excel_file = doc['excel_file']

            # you may put validations here to check extension or file size

            wb = openpyxl.load_workbook(excel_file)

            # getting all sheets
            sheets = wb.sheetnames
            print(sheets)

            # getting a particular sheet
            worksheet = wb["Sheet1"]
            print(worksheet)

            # getting active sheet
            active_sheet = wb.active
            print(active_sheet)

            # reading a cell
            # print(worksheet["B"][0].value)

            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            counter = 0
            table = request.POST['fileuploader']
            if table == 'teacher':
                for row in worksheet.iter_rows():
                    row_data = list()
                    # print(row[0].value)

                    for cell in row:
                        row_data.append(str(cell.value))
                        if (counter == 0):
                            continue
                        t = TeacherInfo.objects.filter(tID=(row[3].value))
                        if (t.count() == 0):
                            try:
                                TeacherInfo.objects.create(
                                    tName=row[0].value,
                                    tInitial=row[1].value,
                                    tDesignation=row[2].value,
                                    tID=row[3].value,
                                    tPhone=row[4].value,
                                    tEmail=row[5].value
                                )
                            except:
                                advisor = ""
                    excel_data.append(row_data)
                    print(row_data)
                    counter += 1
            elif table == 'student':
                for row in worksheet.iter_rows():
                    row_data = list()
                    # print(row[0].value)

                    for cell in row:
                        row_data.append(str(cell.value))
                        if (counter == 0):
                            continue
                        s = StudentInfo.objects.filter(stID=(row[0].value))
                        if (s.count() == 0):
                            try:
                                advisor = TeacherInfo.objects.get(tInitial=row[5].value)
                            except:
                                advisor = ""
                            StudentInfo.objects.create(
                                stID=row[0].value,
                                stName=row[1].value,
                                stGender=row[2].value,
                                stPhone=row[3].value,
                                stEmail=row[4].value,
                                stAdvisor=advisor,
                            )

                    excel_data.append(row_data)
                    print(row_data)
                    counter += 1
            elif table == 'course':
                for row in worksheet.iter_rows():
                    row_data = list()
                    # print(row[0].value)

                    for cell in row:
                        row_data.append(str(cell.value))
                        if (counter == 0):
                            continue
                        c = Courses.objects.filter(courseCode=(row[0].value))
                        if (c.count() == 0):
                            try:
                                Courses.objects.create(
                                    courseCode=row[0].value,
                                    courseTitle=row[1].value,
                                    courseCredit=row[2].value,
                                    level=row[3].value,
                                    term=row[4].value
                                )
                            except:
                                advisor = ""

                    excel_data.append(row_data)
                    print(row_data)
                    counter += 1
            elif table == 'waiver':
                for row in worksheet.iter_rows():
                    row_data = list()
                    # print(row[0].value)

                    for cell in row:
                        row_data.append(str(cell.value))
                        if (counter == 0):
                            continue
                        # try:
                        # print(row[0].value)
                        # student = StudentInfo.objects.get(stID=row[0].value)
                        # except:
                        #     student = ""
                        # try:
                        semester = SemesterInfo.objects.get(semesterCode='201')
                        # except:
                        #     semester = ""
                        c = WaiverInfo.objects.filter(studentID=row[0].value)
                        if (c.count() == 0):
                            try:
                                WaiverInfo.objects.create(
                                    studentID=row[0].value,
                                    semester=semester,
                                    existingWaiver=row[1].value,
                                    specialWaiver=row[2].value,
                                    totalWaiver=row[3].value,
                                    amountofWaiver=row[4].value,
                                    percentofWaiver=row[5].value,
                                )
                            except:
                                msg = ""

                    excel_data.append(row_data)
                    print(row_data)
                    counter += 1

            messages.success(request, f'Successfully Uploaded file into the system!')
            return redirect(index)
    except :
        messages.error(request, f'Something went wrong. Invalid file type!')
        return redirect(index)
        # return redirect('error404','please choose a file!!!')

def error404(request, msg):
    return render(request,'404.html')

def prevPage(request):
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
def reportGenerator(request):
    studentinfo = StudentInfo.objects.all()
    semester=SemesterInfo.objects.get(semesterCode='201')
    regStudents= CoursePreRegistration.objects.filter(semester=semester).values('student_id', 'student__stID','student__stEmail','student__stAdvisor__tInitial','student__stName', 'student__remarks__remark').annotate(credit=Sum('course__courseCredit')).distinct()
    # notregistered = studentinfo.exclude(stID__in=[o['student__stID'] for o in regStudents]).values('id','stID','stName','stPhone', 'stEmail', 'stAdvisor','remarks__remark')
    teacher = TeacherInfo.objects.values('tInitial')
    studentnum = TeacherInfo.objects.all().values('tInitial','tName').annotate(numofstudent=Count('studentinfo',distinct=True))
    finalregistered = TeacherInfo.objects.all().values('tInitial').annotate(
        numoffinalstudent=Count('studentinfo__courseregistration'  , filter=Q(
            studentinfo__courseregistration__registered=True))).order_by('tInitial')
    # finalregistered = TeacherInfo.objects.filter(studentinfo__courseregistration__registered =True).values('tInitial').annotate(numoffinalstudent= Case(When(Count('studentinfo__courseregistration' )=='', then=Value('0')), default=Value(Count('studentinfo__courseregistration' )))).order_by('tInitial')
    # tsummury = CoursePreRegistration.objects.filter(student__stAdvisor__tInitial__in=teacher).values( 'student__stAdvisor__tInitial').annotate(reg=Count('student__stID', distinct=True))\
    tsummury = TeacherInfo.objects.all().values(
        'tInitial').annotate(reg=Count('studentinfo__studentCredits__student', distinct=True))
        # .extra(select={'val': "select count(stID, distinct =TRUE ) from student.studentInfo where stAdvisor = $s"}, select_params=('student__stAdvisor__tInitial',))
    print(tsummury)

    # ,student= teacher.filter(tInitial='student__stAdvisor__tInitial').values('numofstudent') ).distinct()
     # count = CoursePreRegistration.objects.filter(semester=semester, course__courseCode=ccode, section=sec).values(
    #     'id').count()
    csummary1 = CoursePreRegistration.objects.filter(semester=semester).values('course__courseCode', 'section').annotate(nstudent=Count('id')).order_by('course__level', 'course__term', 'course__courseCode')
    csummary= CoursePreRegistration.objects.filter(semester=semester).values('course_id','course__courseCode','course__courseCredit', 'course__courseTitle', 'course__totalSection').annotate(regstudents = Count('course_id')).order_by('course__level', 'course__term', 'course__courseCode')
        # .extra(select={'seca': "select student_id from CoursePreRegistration  where id = 400"}, )
    # # sdetails = StudentInfo.objects.values('stID',
    # #                                        'stName',
    # #                                        'stEmail',
    # #                                        'stPhone',
    # #                                      'stAdvisor__tInitial',
    # #                                      'stAdvisor__tName',
    # #                                        'remarks__remark' ).distinct().order_by('stAdvisor__tInitial')
    sdetails=StudentInfo.objects.raw('''SELECT DISTINCT student_studentinfo.stID,student_studentinfo.id, student_studentinfo.stName,student_studentinfo.stEmail,
student_studentinfo.stPhone,teacher_teacherinfo.tInitial,teacher_teacherinfo.tName, 
adminpanel_remarks.remark,
CASE WHEN adminpanel_coursepreregistration.student_id IS NOT NULL THEN 'Registered' ELSE 'Not Registered' END as Registered
FROM student_studentinfo 
LEFT JOIN users_user on student_studentinfo.stID=users_user.username
LEFT JOIN teacher_teacherinfo on teacher_teacherinfo.id = student_studentinfo.stAdvisor_id
LEFT JOIN adminpanel_coursepreregistration on  student_studentinfo.id = adminpanel_coursepreregistration.student_id
LEFT JOIN adminpanel_remarks on  student_studentinfo.id= adminpanel_remarks.student_id
order by teacher_teacherinfo.tInitial''')
    print(sdetails[0].stName+"   "+ sdetails[1].stName+"   "+ sdetails[2].stName+"   "+ sdetails[3].stName)
    context = {
        'regStudents':regStudents,
        # 'notregistered':notregistered,
        'tsummury':tsummury,
        'studentnum':studentnum,
        'csummary':csummary,
        'csummary1':csummary1,
        'sdetails':sdetails,
        'finalregistered':finalregistered,
    }
    return render(request, 'adminpanel/report.html', context)
@csrf_exempt
def gettakencourses(request):
    cdetails = CoursePreRegistration.objects.filter(semester__semesterCode='201').values('student__stID',
                                                                                         'course__courseCode',
                                                                                         'course__courseTitle',
                                                                                         'course__courseCredit',
                                                                                         'section').distinct().order_by('student__stAdvisor__tInitial')

    data = json.dumps(list(cdetails))
    return HttpResponse(data)

@csrf_exempt
def getcourses(request):
    courses = Courses.objects.values('id','courseCode', 'courseTitle', 'totalSection').order_by('level', 'term', 'courseCode')
    data = json.dumps(list(courses))
    return HttpResponse(data)

@csrf_exempt
def getteacherid(request):
    tinitial = request.POST['tinitial']
    teacher = TeacherInfo.objects.get(tInitial=tinitial)
    data = teacher.id
    return HttpResponse(data)